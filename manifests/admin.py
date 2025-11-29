from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.models import User
from django.utils.html import format_html
from django.db import models
from django.shortcuts import render, redirect
from django.contrib import messages
from django.urls import path
from import_export import resources
from import_export.admin import ImportExportModelAdmin
from import_export.fields import Field
from import_export.widgets import DateWidget
from .models import ManifestEntry, DatabaseYear, ContainerType, Ship, Pavilion, ImportTemplate
from datetime import datetime


class CustomDateWidget(DateWidget):
    """Widget custom pentru format de data DD.MM.YYYY"""
    def clean(self, value, row=None, **kwargs):
        if not value:
            return None

        # Daca este deja un obiect date, returneaza-l
        if isinstance(value, datetime):
            return value.date()

        # Incearca sa parseze data in format DD.MM.YYYY
        if isinstance(value, str):
            for fmt in ['%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d']:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue

        return super().clean(value, row, **kwargs)


class ManifestEntryResource(resources.ModelResource):
    """Resource pentru import/export Excel"""

    # Mapare coloane Excel -> Model Django
    numar_manifest = Field(attribute='numar_manifest', column_name='numar manifest')
    numar_permis = Field(attribute='numar_permis', column_name='numar permis')
    numar_pozitie = Field(attribute='numar_pozitie', column_name='numar pozitie')
    cerere_operatiune = Field(attribute='cerere_operatiune', column_name='cerere operatiune')
    data_inregistrare = Field(attribute='data_inregistrare', column_name='data inregistrare', widget=CustomDateWidget(format='%d.%m.%Y'))
    container = Field(attribute='container', column_name='container')
    numar_colete = Field(attribute='numar_colete', column_name='numar colete')
    greutate_bruta = Field(attribute='greutate_bruta', column_name='greutate bruta')
    descriere_marfa = Field(attribute='descriere_marfa', column_name='descriere marfa')
    tip_operatiune = Field(attribute='tip_operatiune', column_name='tip operatiune')
    nume_nava = Field(attribute='nume_nava', column_name='nume nava')
    pavilion_nava = Field(attribute='pavilion_nava', column_name='pavilion nava')
    numar_sumara = Field(attribute='numar_sumara', column_name='numar sumara')
    tip_container = Field(attribute='tip_container', column_name='tip container')
    linie_maritima = Field(attribute='linie_maritima', column_name='linie maritima')

    class Meta:
        model = ManifestEntry
        skip_unchanged = True
        report_skipped = True
        import_id_fields = ['numar_manifest', 'numar_pozitie', 'container']
        fields = (
            'numar_manifest', 'numar_permis', 'numar_pozitie', 'cerere_operatiune',
            'data_inregistrare', 'container', 'numar_colete', 'greutate_bruta',
            'descriere_marfa', 'tip_operatiune', 'nume_nava', 'pavilion_nava', 'numar_sumara',
            'tip_container', 'linie_maritima', 'model_container'
        )

    def before_import_row(self, row, **kwargs):
        """Genereaza numar_curent automat inainte de import"""
        # Get max numar_curent din baza de date
        max_numar = ManifestEntry.objects.aggregate(models.Max('numar_curent'))['numar_curent__max']
        if max_numar is None:
            max_numar = 0
        # Nu incrementam aici, o facem in after_import_row pentru fiecare rand nou

    def after_import_row(self, row, row_result, **kwargs):
        """Seteaza numar_curent pentru randurile noi sau actualizate"""
        if row_result.import_type in ['new', 'update']:
            instance = row_result.instance
            if instance and instance.pk:
                # Daca e nou, setam numar_curent
                if instance.numar_curent == 0:
                    max_numar = ManifestEntry.objects.aggregate(models.Max('numar_curent'))['numar_curent__max']
                    if max_numar is None:
                        max_numar = 0
                    instance.numar_curent = max_numar + 1
                    instance.save(update_fields=['numar_curent'])


# Admin actions globale - trebuie definite INAINTE de clase
def sync_lookup_tables_action(modeladmin, request, queryset):
    """Actiune admin pentru sincronizarea tabelelor de referinta"""
    from django.core.management import call_command
    from io import StringIO

    # Capteaza output-ul command-ului
    out = StringIO()
    call_command('sync_lookup_tables', stdout=out)
    output = out.getvalue()

    # Afiseaza output-ul ca mesaj de succes
    modeladmin.message_user(request, output)

sync_lookup_tables_action.short_description = 'Sincronizeaza tabele (ContainerType, Ship, Pavilion)'


@admin.register(ManifestEntry)
class ManifestEntryAdmin(ImportExportModelAdmin):
    """Admin pentru ManifestEntry cu import/export Excel"""

    resource_class = ManifestEntryResource
    actions = [sync_lookup_tables_action]
    change_list_template = 'admin/manifests/manifestentry_changelist.html'

    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }
        js = ('admin/js/resizable_columns.js', 'admin/js/move_filters.js', 'admin/js/search_reset.js', 'admin/js/highlight_duplicates.js')

    list_display = [
        'format_numar_curent',
        'format_numar_manifest', 'format_numar_permis', 'format_numar_pozitie', 'format_cerere_operatiune',
        'format_data_inregistrare', 'format_container', 'format_model_container', 'format_tip_container',
        'format_numar_colete', 'format_greutate_bruta', 'format_descriere_marfa', 'format_tip_operatiune',
        'format_nume_nava', 'format_pavilion_nava', 'format_numar_sumara', 'format_linie_maritima', 'format_observatii'
    ]

    def format_numar_curent(self, obj):
        return obj.numar_curent
    format_numar_curent.short_description = format_html('Nr.<br>Crt.')
    format_numar_curent.admin_order_field = 'numar_curent'

    def format_data_inregistrare(self, obj):
        """Formatare data in DD.MM.YYYY"""
        if obj.data_inregistrare:
            return obj.data_inregistrare.strftime('%d.%m.%Y')
        return '-'
    format_data_inregistrare.short_description = format_html('Data<br>Inregistrare')
    format_data_inregistrare.admin_order_field = 'data_inregistrare'

    def format_container(self, obj):
        """Evidentiaza containerele duplicate cu galben"""
        if not obj.container:
            return '-'

        # Numără câte înregistrări au același container în toată baza de date
        count = ManifestEntry.objects.filter(
            container=obj.container,
            database_year=obj.database_year
        ).count()

        if count > 1:
            # Container duplicat - evidențiază cu galben
            return format_html(
                '<span style="background-color: #fff3cd;" '
                'title="Container duplicat (apare de {} ori în baza de date)">{}</span>',
                count,
                obj.container
            )
        else:
            return obj.container

    format_container.short_description = 'Container'
    format_container.admin_order_field = 'container'

    def format_numar_manifest(self, obj):
        return obj.numar_manifest
    format_numar_manifest.short_description = format_html('Numar<br>Manifest')
    format_numar_manifest.admin_order_field = 'numar_manifest'

    def format_numar_permis(self, obj):
        return obj.numar_permis
    format_numar_permis.short_description = format_html('Numar<br>Permis')
    format_numar_permis.admin_order_field = 'numar_permis'

    def format_numar_pozitie(self, obj):
        return obj.numar_pozitie
    format_numar_pozitie.short_description = format_html('Numar<br>Pozitie')
    format_numar_pozitie.admin_order_field = 'numar_pozitie'

    def format_cerere_operatiune(self, obj):
        return obj.cerere_operatiune
    format_cerere_operatiune.short_description = format_html('Cerere<br>Operatiune')
    format_cerere_operatiune.admin_order_field = 'cerere_operatiune'

    def format_model_container(self, obj):
        return obj.model_container
    format_model_container.short_description = format_html('Model<br>Container')
    format_model_container.admin_order_field = 'model_container'

    def format_tip_container(self, obj):
        return obj.tip_container
    format_tip_container.short_description = format_html('Tip<br>Container')
    format_tip_container.admin_order_field = 'tip_container'

    def format_numar_colete(self, obj):
        return obj.numar_colete
    format_numar_colete.short_description = format_html('Numar<br>Colete')
    format_numar_colete.admin_order_field = 'numar_colete'

    def format_greutate_bruta(self, obj):
        return obj.greutate_bruta
    format_greutate_bruta.short_description = format_html('Greutate<br>Bruta')
    format_greutate_bruta.admin_order_field = 'greutate_bruta'

    def format_descriere_marfa(self, obj):
        return obj.descriere_marfa[:50] + '...' if len(obj.descriere_marfa or '') > 50 else obj.descriere_marfa
    format_descriere_marfa.short_description = format_html('Descriere<br>Marfa')
    format_descriere_marfa.admin_order_field = 'descriere_marfa'

    def format_tip_operatiune(self, obj):
        return obj.tip_operatiune
    format_tip_operatiune.short_description = format_html('Tip<br>Operatiune')
    format_tip_operatiune.admin_order_field = 'tip_operatiune'

    def format_nume_nava(self, obj):
        return obj.nume_nava
    format_nume_nava.short_description = format_html('Nume<br>Nava')
    format_nume_nava.admin_order_field = 'nume_nava'

    def format_pavilion_nava(self, obj):
        return obj.pavilion_nava
    format_pavilion_nava.short_description = format_html('Pavilion<br>Nava')
    format_pavilion_nava.admin_order_field = 'pavilion_nava'

    def format_numar_sumara(self, obj):
        return obj.numar_sumara
    format_numar_sumara.short_description = format_html('Numar<br>Sumara')
    format_numar_sumara.admin_order_field = 'numar_sumara'

    def format_linie_maritima(self, obj):
        return obj.linie_maritima
    format_linie_maritima.short_description = format_html('Linie<br>Maritima')
    format_linie_maritima.admin_order_field = 'linie_maritima'

    def format_observatii(self, obj):
        if obj.observatii:
            lungime = len(obj.observatii)
            if lungime > 50:
                display_text = obj.observatii[:50] + '...'
            else:
                display_text = obj.observatii
            return format_html('<span data-obs-length="{}">{}</span>', lungime, display_text)
        return '-'
    format_observatii.short_description = 'Observatii'
    format_observatii.admin_order_field = 'observatii'

    list_filter = [
        'data_inregistrare', 'numar_manifest', 'linie_maritima'
    ]

    search_fields = [
        'numar_manifest', 'container', 'model_container', 'nume_nava',
        'numar_permis', 'descriere_marfa', 'linie_maritima'
    ]

    readonly_fields = ['numar_curent', 'model_container', 'created_at', 'updated_at']

    fieldsets = (
        ('Informatii Principale', {
            'fields': ('numar_curent', 'numar_manifest', 'numar_permis', 'numar_pozitie', 'cerere_operatiune', 'data_inregistrare')
        }),
        ('Container', {
            'fields': ('container', 'tip_container', 'model_container', 'numar_colete', 'greutate_bruta')
        }),
        ('Marfa', {
            'fields': ('descriere_marfa', 'tip_operatiune')
        }),
        ('Nava', {
            'fields': ('nume_nava', 'pavilion_nava', 'numar_sumara', 'linie_maritima')
        }),
        ('Observatii', {
            'fields': ('observatii',)
        }),
        ('Metadata', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    date_hierarchy = 'data_inregistrare'
    list_per_page = 50

    def get_urls(self):
        """Adaugă URL-uri custom pentru import personalizat"""
        urls = super().get_urls()
        custom_urls = [
            path('import-personalizat/', self.admin_site.admin_view(self.custom_import_view), name='manifests_manifestentry_custom_import'),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        """Adaugă buton pentru import personalizat"""
        extra_context = extra_context or {}
        extra_context['show_custom_import_button'] = True
        return super().changelist_view(request, extra_context)

    def custom_import_view(self, request):
        """View pentru import personalizat cu template și mapare coloane - cu preview"""
        from django.shortcuts import render
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        import openpyxl
        import xlrd
        from decimal import Decimal
        import json

        # Verifică dacă e confirmarea importului (Step 2)
        if request.method == 'POST' and request.POST.get('confirm_import') == 'true':
            try:
                # Citește datele din sesiune
                preview_data = request.session.get('import_preview_data')
                if not preview_data:
                    messages.error(request, 'Sesiunea a expirat. Vă rugăm să reîncărcați fișierul.')
                    return HttpResponseRedirect(reverse('admin:manifests_manifestentry_custom_import'))

                active_year_id = preview_data.get('active_year_id')
                entries_data = preview_data.get('entries_data', [])

                active_year = DatabaseYear.objects.get(id=active_year_id)
                entries_created = 0

                # Creează înregistrările în baza de date
                for entry_dict in entries_data:
                    # Convertește înapoi data dacă există
                    if 'data_inregistrare' in entry_dict and entry_dict['data_inregistrare']:
                        from datetime import datetime
                        entry_dict['data_inregistrare'] = datetime.strptime(entry_dict['data_inregistrare'], '%Y-%m-%d').date()

                    # Convertește Decimal pentru greutate_bruta
                    if 'greutate_bruta' in entry_dict and entry_dict['greutate_bruta']:
                        entry_dict['greutate_bruta'] = Decimal(str(entry_dict['greutate_bruta']))

                    entry_dict['database_year'] = active_year

                    # Calculează numar_curent automat
                    max_curent = ManifestEntry.objects.filter(database_year=active_year).aggregate(models.Max('numar_curent'))['numar_curent__max'] or 0
                    entry_dict['numar_curent'] = max_curent + 1

                    ManifestEntry.objects.create(**entry_dict)
                    entries_created += 1

                # Șterge datele din sesiune
                del request.session['import_preview_data']
                request.session.modified = True

                messages.success(request, f'Import finalizat cu succes! {entries_created} înregistrări au fost create.')
                return HttpResponseRedirect(reverse('admin:manifests_manifestentry_changelist'))

            except Exception as e:
                messages.error(request, f'Eroare la salvarea datelor: {str(e)}')
                return HttpResponseRedirect(reverse('admin:manifests_manifestentry_custom_import'))

        # Procesare upload și preview (Step 1)
        if request.method == 'POST':
            # Procesare import
            template_id = request.POST.get('template_id')
            file = request.FILES.get('file')

            # Citește câmpurile manuale
            manual_numar_manifest = request.POST.get('numar_manifest', '').strip()
            manual_numar_permis = request.POST.get('numar_permis', '').strip()
            manual_data_inregistrare = request.POST.get('data_inregistrare', '').strip()
            manual_cerere_operatiune = request.POST.get('cerere_operatiune', '').strip()
            manual_nume_nava = request.POST.get('nume_nava', '').strip()
            manual_pavilion_nava = request.POST.get('pavilion_nava', '').strip()

            if not template_id or not file:
                messages.error(request, 'Vă rugăm să selectați un template și să încărcați un fișier.')
                return HttpResponseRedirect(reverse('admin:manifests_manifestentry_custom_import'))

            if not manual_numar_manifest:
                messages.error(request, 'Numărul manifestului este obligatoriu.')
                return HttpResponseRedirect(reverse('admin:manifests_manifestentry_custom_import'))

            try:
                template = ImportTemplate.objects.get(id=template_id)
                active_year = DatabaseYear.objects.filter(is_active=True).first()

                if not active_year:
                    messages.error(request, 'Nu există un an activ selectat. Vă rugăm să activați un an din secțiunea "Ani Baze Date".')
                    return HttpResponseRedirect(reverse('admin:manifests_manifestentry_custom_import'))

                # Procesare fișier și colectare date pentru preview
                preview_entries = []
                errors = []

                # Funcție helper pentru a converti literă coloană Excel în index (A=1, B=2, etc.)
                def excel_col_to_index(col_letter):
                    """Convertește literă coloană Excel (A, B, AA, etc.) în index numeric"""
                    col_letter = col_letter.upper().strip()
                    result = 0
                    for char in col_letter:
                        result = result * 26 + (ord(char) - ord('A') + 1)
                    return result

                # Detectează formatul fișierului
                file_extension = file.name.split('.')[-1].lower()

                if file_extension == 'xlsx' or template.format_fisier == 'xlsx':
                    workbook = openpyxl.load_workbook(file)
                    sheet = workbook.active

                    # Procesare rânduri
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=template.rand_start), template.rand_start):
                        try:
                            entry_data = {}

                            # Mapare coloane din template (excel_col_letter este A, B, C, etc.)
                            for db_field, excel_col_letter in template.mapare_coloane.items():
                                # Sari peste câmpurile manuale - acestea nu se iau din Excel
                                if db_field in ['numar_manifest', 'numar_permis', 'data_inregistrare', 'cerere_operatiune', 'nume_nava', 'pavilion_nava']:
                                    continue

                                try:
                                    col_idx = excel_col_to_index(excel_col_letter)
                                    cell_value = row[col_idx - 1].value

                                    if cell_value is not None:
                                        # Conversii speciale pentru tipuri de date
                                        if db_field in ['numar_colete', 'numar_pozitie'] and cell_value:
                                            entry_data[db_field] = int(float(cell_value)) if cell_value else None
                                        elif db_field == 'greutate_bruta' and cell_value:
                                            entry_data[db_field] = Decimal(str(cell_value)) if cell_value else None
                                        elif db_field == 'tip_operatiune':
                                            # Conversie tip operațiune: IMP -> I, TRS -> T
                                            tip_op = str(cell_value).strip().upper()
                                            if tip_op == 'IMP':
                                                entry_data[db_field] = 'I'
                                            elif tip_op == 'TRS':
                                                entry_data[db_field] = 'T'
                                            elif tip_op in ['I', 'T']:
                                                entry_data[db_field] = tip_op
                                            else:
                                                # Validare: doar I sau T sunt permise
                                                raise ValueError(f"Tip operațiune invalid: '{tip_op}'. Doar I, T, IMP sau TRS sunt permise.")
                                        else:
                                            entry_data[db_field] = str(cell_value).strip() if cell_value else ''
                                except (IndexError, ValueError) as e:
                                    if 'Tip operațiune invalid' in str(e):
                                        raise  # Propagă eroarea de validare
                                    continue

                            # Aplică câmpurile manuale la toate rândurile
                            entry_data['numar_manifest'] = manual_numar_manifest
                            entry_data['numar_permis'] = manual_numar_permis
                            entry_data['cerere_operatiune'] = manual_cerere_operatiune
                            if manual_nume_nava:
                                entry_data['nume_nava'] = manual_nume_nava
                            if manual_pavilion_nava:
                                entry_data['pavilion_nava'] = manual_pavilion_nava

                            # Conversie data din format YYYY-MM-DD (HTML date input)
                            if manual_data_inregistrare:
                                from datetime import datetime
                                try:
                                    entry_data['data_inregistrare'] = datetime.strptime(manual_data_inregistrare, '%Y-%m-%d').date()
                                except ValueError:
                                    pass  # Dacă data nu e validă, nu o setează

                            if entry_data:
                                # Convertește date și Decimal în string pentru sesiune
                                entry_data_serializable = {}
                                for key, value in entry_data.items():
                                    if hasattr(value, 'isoformat'):  # datetime.date
                                        entry_data_serializable[key] = value.isoformat()
                                    elif isinstance(value, Decimal):
                                        entry_data_serializable[key] = str(value)
                                    else:
                                        entry_data_serializable[key] = value

                                preview_entries.append(entry_data_serializable)

                        except Exception as e:
                            errors.append(f"Rând {row_idx}: {str(e)}")

                elif file_extension == 'xls' or template.format_fisier == 'xls':
                    # Încearcă să citești cu xlrd (pentru fișiere .xls vechi)
                    xlrd_success = False
                    try:
                        workbook = xlrd.open_workbook(file_contents=file.read())
                        sheet = workbook.sheet_by_index(0)
                        xlrd_success = True
                    except Exception as xlrd_error:
                        # Fallback la openpyxl pentru fișiere .xls cu format nou
                        try:
                            file.seek(0)  # Resetează pointer-ul fișierului
                            workbook = openpyxl.load_workbook(file)
                            sheet = workbook.active
                            xlrd_success = False  # Folosim openpyxl
                        except Exception as openpyxl_error:
                            raise Exception(f"Nu s-a putut citi fișierul .xls. Eroare xlrd: {xlrd_error}. Eroare openpyxl: {openpyxl_error}")

                    # Procesare cu xlrd sau openpyxl în funcție de ce a funcționat
                    if xlrd_success:
                        # Procesare cu xlrd (format vechi .xls)
                        for row_idx in range(template.rand_start - 1, sheet.nrows):
                            try:
                                entry_data = {}

                                for db_field, excel_col_letter in template.mapare_coloane.items():
                                    # Sari peste câmpurile manuale - acestea nu se iau din Excel
                                    if db_field in ['numar_manifest', 'numar_permis', 'data_inregistrare', 'cerere_operatiune', 'nume_nava', 'pavilion_nava']:
                                        continue

                                    try:
                                        col_idx = excel_col_to_index(excel_col_letter) - 1  # XLS folosește index 0-based
                                        cell_value = sheet.cell_value(row_idx, col_idx)

                                        if cell_value:
                                            if db_field in ['numar_colete', 'numar_pozitie']:
                                                entry_data[db_field] = int(float(cell_value)) if cell_value else None
                                            elif db_field == 'greutate_bruta':
                                                entry_data[db_field] = Decimal(str(cell_value)) if cell_value else None
                                            elif db_field == 'tip_operatiune':
                                                # Conversie tip operațiune: IMP -> I, TRS -> T
                                                tip_op = str(cell_value).strip().upper()
                                                if tip_op == 'IMP':
                                                    entry_data[db_field] = 'I'
                                                elif tip_op == 'TRS':
                                                    entry_data[db_field] = 'T'
                                                elif tip_op in ['I', 'T']:
                                                    entry_data[db_field] = tip_op
                                                else:
                                                    # Validare: doar I sau T sunt permise
                                                    raise ValueError(f"Tip operațiune invalid: '{tip_op}'. Doar I, T, IMP sau TRS sunt permise.")
                                            else:
                                                entry_data[db_field] = str(cell_value).strip()
                                    except (IndexError, ValueError) as e:
                                        if 'Tip operațiune invalid' in str(e):
                                            raise  # Propagă eroarea de validare
                                        continue

                                # Aplică câmpurile manuale la toate rândurile
                                entry_data['numar_manifest'] = manual_numar_manifest
                                entry_data['numar_permis'] = manual_numar_permis
                                entry_data['cerere_operatiune'] = manual_cerere_operatiune
                                if manual_nume_nava:
                                    entry_data['nume_nava'] = manual_nume_nava
                                if manual_pavilion_nava:
                                    entry_data['pavilion_nava'] = manual_pavilion_nava

                                # Conversie data din format YYYY-MM-DD (HTML date input)
                                if manual_data_inregistrare:
                                    from datetime import datetime
                                    try:
                                        entry_data['data_inregistrare'] = datetime.strptime(manual_data_inregistrare, '%Y-%m-%d').date()
                                    except ValueError:
                                        pass  # Dacă data nu e validă, nu o setează

                                if entry_data:
                                    # Convertește date și Decimal în string pentru sesiune
                                    entry_data_serializable = {}
                                    for key, value in entry_data.items():
                                        if hasattr(value, 'isoformat'):  # datetime.date
                                            entry_data_serializable[key] = value.isoformat()
                                        elif isinstance(value, Decimal):
                                            entry_data_serializable[key] = str(value)
                                        else:
                                            entry_data_serializable[key] = value

                                    preview_entries.append(entry_data_serializable)

                            except Exception as e:
                                errors.append(f"Rând {row_idx + 1}: {str(e)}")

                    else:
                        # Procesare cu openpyxl (fallback pentru .xls cu format nou)
                        for row_idx, row in enumerate(sheet.iter_rows(min_row=template.rand_start), template.rand_start):
                            try:
                                entry_data = {}

                                for db_field, excel_col_letter in template.mapare_coloane.items():
                                    if db_field in ['numar_manifest', 'numar_permis', 'data_inregistrare', 'cerere_operatiune', 'nume_nava', 'pavilion_nava']:
                                        continue

                                    try:
                                        col_idx = excel_col_to_index(excel_col_letter)
                                        cell_value = row[col_idx - 1].value

                                        if cell_value is not None:
                                            if db_field in ['numar_colete', 'numar_pozitie'] and cell_value:
                                                entry_data[db_field] = int(float(cell_value)) if cell_value else None
                                            elif db_field == 'greutate_bruta' and cell_value:
                                                entry_data[db_field] = Decimal(str(cell_value)) if cell_value else None
                                            elif db_field == 'tip_operatiune':
                                                tip_op = str(cell_value).strip().upper()
                                                if tip_op == 'IMP':
                                                    entry_data[db_field] = 'I'
                                                elif tip_op == 'TRS':
                                                    entry_data[db_field] = 'T'
                                                elif tip_op in ['I', 'T']:
                                                    entry_data[db_field] = tip_op
                                                else:
                                                    raise ValueError(f"Tip operațiune invalid: '{tip_op}'. Doar I, T, IMP sau TRS sunt permise.")
                                            else:
                                                entry_data[db_field] = str(cell_value).strip() if cell_value else ''
                                    except (IndexError, ValueError) as e:
                                        if 'Tip operațiune invalid' in str(e):
                                            raise
                                        continue

                                # Aplică câmpurile manuale
                                entry_data['numar_manifest'] = manual_numar_manifest
                                entry_data['numar_permis'] = manual_numar_permis
                                entry_data['cerere_operatiune'] = manual_cerere_operatiune
                                if manual_nume_nava:
                                    entry_data['nume_nava'] = manual_nume_nava
                                if manual_pavilion_nava:
                                    entry_data['pavilion_nava'] = manual_pavilion_nava

                                if manual_data_inregistrare:
                                    from datetime import datetime
                                    try:
                                        entry_data['data_inregistrare'] = datetime.strptime(manual_data_inregistrare, '%Y-%m-%d').date()
                                    except ValueError:
                                        pass

                                if entry_data:
                                    entry_data_serializable = {}
                                    for key, value in entry_data.items():
                                        if hasattr(value, 'isoformat'):
                                            entry_data_serializable[key] = value.isoformat()
                                        elif isinstance(value, Decimal):
                                            entry_data_serializable[key] = str(value)
                                        else:
                                            entry_data_serializable[key] = value

                                    preview_entries.append(entry_data_serializable)

                            except Exception as e:
                                errors.append(f"Rând {row_idx}: {str(e)}")

                else:
                    # Format de fișier nerecunoscut
                    messages.error(request, f'Format de fișier nerecunoscut: .{file_extension}. Vă rugăm să încărcați un fișier .xlsx sau .xls')
                    return HttpResponseRedirect(reverse('admin:manifests_manifestentry_custom_import'))

                # Salvează datele în sesiune pentru preview
                if errors:
                    messages.error(request, f'Erori la procesarea fișierului: {"; ".join(errors[:5])}')
                    return HttpResponseRedirect(reverse('admin:manifests_manifestentry_custom_import'))

                if not preview_entries:
                    messages.warning(request, 'Nu s-au găsit date de importat în fișier.')
                    return HttpResponseRedirect(reverse('admin:manifests_manifestentry_custom_import'))

                # Salvează în sesiune pentru Step 2
                request.session['import_preview_data'] = {
                    'active_year_id': active_year.id,
                    'entries_data': preview_entries,
                    'template_name': template.nume,
                    'manual_fields': {
                        'numar_manifest': manual_numar_manifest,
                        'numar_permis': manual_numar_permis,
                        'data_inregistrare': manual_data_inregistrare,
                        'cerere_operatiune': manual_cerere_operatiune,
                        'nume_nava': manual_nume_nava,
                        'pavilion_nava': manual_pavilion_nava,
                    }
                }
                request.session.modified = True

                # Redirecționează la preview
                context = {
                    **self.admin_site.each_context(request),
                    'title': 'Previzualizare Import',
                    'preview_entries': preview_entries[:50],  # Primele 50 pentru afișare
                    'total_entries': len(preview_entries),
                    'template_name': template.nume,
                    'manual_fields': {
                        'numar_manifest': manual_numar_manifest,
                        'numar_permis': manual_numar_permis,
                        'data_inregistrare': manual_data_inregistrare,
                        'cerere_operatiune': manual_cerere_operatiune,
                        'nume_nava': manual_nume_nava,
                        'pavilion_nava': manual_pavilion_nava,
                    },
                    'opts': self.model._meta,
                }
                return render(request, 'admin/manifests/import_preview.html', context)

            except ImportTemplate.DoesNotExist:
                messages.error(request, 'Template-ul selectat nu există.')
            except Exception as e:
                messages.error(request, f'Eroare la import: {str(e)}')

            return HttpResponseRedirect(reverse('admin:manifests_manifestentry_custom_import'))

        # GET request - afișează formularul
        templates = ImportTemplate.objects.all()
        context = {
            **self.admin_site.each_context(request),
            'title': 'Import Personalizat Manifest',
            'templates': templates,
            'opts': self.model._meta,
        }
        return render(request, 'admin/manifests/custom_import.html', context)


# Admin pentru DatabaseYear
@admin.register(DatabaseYear)
class DatabaseYearAdmin(admin.ModelAdmin):
    """Admin pentru gestionarea bazelor de date pe ani"""

    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }
        js = ('admin/js/resizable_columns.js', 'admin/js/move_filters.js', 'admin/js/search_reset.js', 'admin/js/highlight_duplicates.js')

    list_display = ['year', 'is_active', 'entries_count', 'created_at']
    list_filter = ['is_active', 'year']
    search_fields = ['year']
    actions = ['activate_year']

    def entries_count(self, obj):
        return obj.entries.count()
    entries_count.short_description = 'Numar Intrari'

    def activate_year(self, request, queryset):
        # Dezactiveaza toate
        DatabaseYear.objects.all().update(is_active=False)
        # Activeaza selectiile
        count = queryset.update(is_active=True)
        self.message_user(request, f'{count} an(i) activat(i) cu succes.')
    activate_year.short_description = 'Activeaza anul selectat'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('create-new-year/', self.admin_site.admin_view(self.create_new_year_view), name='create_new_year'),
        ]
        return custom_urls + urls

    def create_new_year_view(self, request):
        """View pentru crearea unui an nou"""
        if request.method == 'POST':
            year = request.POST.get('year')
            try:
                year = int(year)
                # Verifica daca exista deja
                if DatabaseYear.objects.filter(year=year).exists():
                    messages.error(request, f'Anul {year} exista deja in baza de date.')
                else:
                    # Creaza anul nou
                    new_year = DatabaseYear.objects.create(year=year, is_active=False)
                    messages.success(request, f'Anul {year} a fost creat cu succes! Poti acum importa date pentru acest an.')
                return redirect('..')
            except ValueError:
                messages.error(request, 'Anul trebuie sa fie un numar valid.')
                return redirect('..')

        # GET request - afiseaza formular
        from django.template.response import TemplateResponse
        context = {
            **self.admin_site.each_context(request),
            'title': 'Creare An Nou',
            'current_year': datetime.now().year,
            'next_year': datetime.now().year + 1,
        }
        return TemplateResponse(request, 'admin/create_new_year.html', context)

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_create_year_button'] = True
        return super().changelist_view(request, extra_context=extra_context)


# Admin pentru Pavilion
@admin.register(Pavilion)
class PavilionAdmin(admin.ModelAdmin):
    """Admin pentru pavilioane"""

    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }
        js = ('admin/js/resizable_columns.js', 'admin/js/move_filters.js', 'admin/js/search_reset.js', 'admin/js/highlight_duplicates.js')

    list_display = ['nume', 'nume_tara', 'preview_imagine', 'ships_count', 'created_at']
    search_fields = ['nume', 'nume_tara']
    readonly_fields = ['preview_imagine_large', 'created_at', 'updated_at']

    fieldsets = (
        ('Informatii Pavilion', {
            'fields': ('nume', 'nume_tara', 'imagine', 'preview_imagine_large')
        }),
        ('Metadata', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def preview_imagine(self, obj):
        if obj.imagine:
            return format_html('<img src="{}" width="50" height="30" style="object-fit: contain;" />', obj.imagine.url)
        return '-'
    preview_imagine.short_description = 'Preview'

    def preview_imagine_large(self, obj):
        if obj.imagine:
            return format_html('<img src="{}" width="200" style="object-fit: contain;" />', obj.imagine.url)
        return '-'
    preview_imagine_large.short_description = 'Preview Imagine'

    def ships_count(self, obj):
        return obj.ships.count()
    ships_count.short_description = 'Numar Nave'


# Admin pentru Ship
@admin.register(Ship)
class ShipAdmin(admin.ModelAdmin):
    """Admin pentru nave"""

    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }
        js = ('admin/js/resizable_columns.js', 'admin/js/move_filters.js', 'admin/js/search_reset.js', 'admin/js/highlight_duplicates.js')

    list_display = ['nume', 'linie_maritima', 'pavilion', 'preview_imagine', 'preview_pavilion', 'entries_count', 'created_at']
    list_filter = ['linie_maritima', 'pavilion']
    search_fields = ['nume', 'linie_maritima']
    readonly_fields = ['preview_imagine_large', 'created_at', 'updated_at']
    autocomplete_fields = ['pavilion']

    fieldsets = (
        ('Informatii Nava', {
            'fields': ('nume', 'linie_maritima', 'pavilion', 'imagine', 'preview_imagine_large', 'descriere')
        }),
        ('Metadata', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def preview_imagine(self, obj):
        if obj.imagine:
            return format_html('<img src="{}" width="60" height="40" style="object-fit: cover; border-radius: 4px;" />', obj.imagine.url)
        return '-'
    preview_imagine.short_description = 'Nava'

    def preview_imagine_large(self, obj):
        if obj.imagine:
            return format_html('<img src="{}" width="300" style="object-fit: contain;" />', obj.imagine.url)
        return '-'
    preview_imagine_large.short_description = 'Preview Imagine'

    def preview_pavilion(self, obj):
        if obj.pavilion and obj.pavilion.imagine:
            return format_html('<img src="{}" width="50" height="30" style="object-fit: contain;" />', obj.pavilion.imagine.url)
        return '-'
    preview_pavilion.short_description = 'Pavilion'

    def entries_count(self, obj):
        return obj.entries.count()
    entries_count.short_description = 'Intrari'


# Filtru custom pentru containere cu/fara imagine
class HasImageFilter(admin.SimpleListFilter):
    """Filtru pentru containere cu sau fara imagine"""
    title = 'are imagine'
    parameter_name = 'has_image'

    def lookups(self, request, model_admin):
        return (
            ('yes', 'Cu imagine'),
            ('no', 'Fără imagine'),
        )

    def queryset(self, request, queryset):
        if self.value() == 'yes':
            return queryset.exclude(imagine='')
        if self.value() == 'no':
            return queryset.filter(imagine='')
        return queryset


# Admin pentru ContainerType
@admin.register(ContainerType)
class ContainerTypeAdmin(admin.ModelAdmin):
    """Admin pentru tipuri containere cu extragere automata din ManifestEntry"""

    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }
        js = ('admin/js/resizable_columns.js', 'admin/js/move_filters.js', 'admin/js/search_reset.js', 'admin/js/highlight_duplicates.js')

    list_display = ['model_container', 'tip_container', 'preview_imagine', 'entries_count', 'created_at']
    list_filter = ['tip_container', HasImageFilter]
    search_fields = ['model_container', 'tip_container']
    readonly_fields = ['preview_imagine_large', 'created_at', 'updated_at']
    actions = ['extract_unique_containers']

    fieldsets = (
        ('Informatii Container', {
            'fields': ('model_container', 'tip_container', 'imagine', 'preview_imagine_large', 'descriere')
        }),
        ('Metadata', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def preview_imagine(self, obj):
        if obj.imagine:
            return format_html('<img src="{}" width="60" height="40" style="object-fit: cover; border-radius: 4px;" />', obj.imagine.url)
        return '-'
    preview_imagine.short_description = 'Preview'

    def preview_imagine_large(self, obj):
        if obj.imagine:
            return format_html('<img src="{}" width="300" style="object-fit: contain;" />', obj.imagine.url)
        return '-'
    preview_imagine_large.short_description = 'Preview Imagine'

    def entries_count(self, obj):
        return obj.entries.count()
    entries_count.short_description = 'Intrari'

    def extract_unique_containers(self, request, queryset):
        """Extrage automat model_container unice din ManifestEntry si le adauga in ContainerType"""
        unique_containers = ManifestEntry.objects.exclude(model_container='').values('model_container', 'tip_container').distinct()

        created_count = 0
        for item in unique_containers:
            model = item['model_container']
            tip = item['tip_container'] or ''
            if model:
                container_type, created = ContainerType.objects.get_or_create(
                    model_container=model,
                    defaults={'tip_container': tip}
                )
                if created:
                    created_count += 1

        self.message_user(request, f'{created_count} tipuri de containere noi au fost create automat.')
    extract_unique_containers.short_description = 'Extrage containere unice din manifeste'


# Customizare User Admin pentru renumire campuri
class CustomUserAdmin(BaseUserAdmin):
    """Admin personalizat pentru utilizatori cu campuri redenumite"""

    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }
        js = ('admin/js/resizable_columns.js', 'admin/js/move_filters.js', 'admin/js/search_reset.js', 'admin/js/highlight_duplicates.js')

    # Customizam fieldsets pentru a redenumi campurile
    fieldsets = (
        (None, {'fields': ('username', 'password')}),
        ('Informații personale', {'fields': ('first_name', 'last_name', 'email')}),
        ('Permisiuni', {
            'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions'),
        }),
        ('Date importante', {'fields': ('last_login', 'date_joined')}),
    )

    # Customizam list_display pentru renumirea campurilor
    list_display = ('username', 'email', 'get_full_name_custom', 'get_company_name', 'is_staff')

    def get_full_name_custom(self, obj):
        """Afiseaza first_name ca 'Nume si prenume'"""
        return obj.first_name or '-'
    get_full_name_custom.short_description = 'Nume si prenume'

    def get_company_name(self, obj):
        """Afiseaza last_name ca 'Nume companie'"""
        return obj.last_name or '-'
    get_company_name.short_description = 'Nume companie'

    # Suprascriu get_form pentru a customiza label-urile campurilor
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if 'first_name' in form.base_fields:
            form.base_fields['first_name'].label = 'Nume si prenume'
        if 'last_name' in form.base_fields:
            form.base_fields['last_name'].label = 'Nume companie'
        return form


# Inregistrare User Admin personalizat
admin.site.unregister(User)
admin.site.register(User, CustomUserAdmin)


# Admin pentru ImportTemplate
@admin.register(ImportTemplate)
class ImportTemplateAdmin(admin.ModelAdmin):
    """Admin pentru gestionarea template-urilor de import"""

    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }
        js = ('admin/js/resizable_columns.js', 'admin/js/move_filters.js', 'admin/js/search_reset.js', 'admin/js/highlight_duplicates.js')

    list_display = ['nume', 'format_fisier', 'rand_start', 'created_at']
    search_fields = ['nume', 'descriere']
    readonly_fields = ['created_at', 'updated_at']

    def get_urls(self):
        """Adaugă URL pentru editare vizuală mapare"""
        urls = super().get_urls()
        from django.urls import path
        custom_urls = [
            path('<path:object_id>/mapare/', self.admin_site.admin_view(self.mapare_view), name='manifests_importtemplate_mapare'),
        ]
        return custom_urls + urls

    def mapare_view(self, request, object_id):
        """View pentru editare vizuală a mapării coloanelor"""
        template = ImportTemplate.objects.get(pk=object_id)

        # Lista coloanelor din ManifestEntry (fără câmpurile manuale)
        # Câmpurile manuale (numar_manifest, numar_permis, data_inregistrare, cerere_operatiune, nume_nava, pavilion_nava)
        # se introduc manual la import, nu se mapează din Excel
        db_fields = [
            ('numar_pozitie', 'Numar Pozitie'),
            ('container', 'Container'),
            ('numar_colete', 'Numar Colete'),
            ('greutate_bruta', 'Greutate Bruta (kg)'),
            ('descriere_marfa', 'Descriere Marfa'),
            ('tip_operatiune', 'Tip Operatiune (I sau T)'),
            ('numar_sumara', 'Numar Sumara'),
            ('tip_container', 'Tip Container'),
            ('linie_maritima', 'Linie Maritima'),
        ]

        if request.method == 'POST':
            # Construiește maparea din formular
            mapare = {}
            for field_name, _ in db_fields:
                excel_col = request.POST.get(f'excel_{field_name}', '').strip()
                if excel_col:
                    mapare[field_name] = excel_col

            template.mapare_coloane = mapare
            template.save()
            messages.success(request, f'Maparea coloanelor pentru template "{template.nume}" a fost salvată cu succes!')
            return redirect('admin:manifests_importtemplate_changelist')

        # Pregătește datele pentru template
        mapping_data = []
        for field_name, field_label in db_fields:
            mapping_data.append({
                'field_name': field_name,
                'field_label': field_label,
                'excel_column': template.mapare_coloane.get(field_name, '')
            })

        context = {
            **self.admin_site.each_context(request),
            'title': f'Mapare Coloane - {template.nume}',
            'template': template,
            'mapping_data': mapping_data,
            'opts': self.model._meta,
        }
        return render(request, 'admin/manifests/template_mapare.html', context)

    def change_view(self, request, object_id, form_url='', extra_context=None):
        """Adaugă buton pentru mapare în pagina de editare"""
        extra_context = extra_context or {}
        extra_context['show_mapare_button'] = True
        return super().change_view(request, object_id, form_url, extra_context)

    fieldsets = (
        ('Informatii Template', {
            'fields': ('nume', 'descriere', 'format_fisier', 'rand_start')
        }),
        ('Metadata', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )


# Customizare titluri admin
admin.site.site_header = "Registru import RE1"
admin.site.site_title = "Registru import RE1"
admin.site.index_title = "Administrare Registru Import"
